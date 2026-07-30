from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent.parent
TARGET = ROOT / "data" / "wechat_messages_template.xlsx"
HEADERS = (
    "消息键",
    "本地时间",
    "会话ID",
    "方向",
    "发送者微信名（账号）",
    "消息类型",
    "消息内容",
    "图片预览",
    "媒体路径",
    "本地消息ID",
    "服务器消息ID",
    "排序序号",
    "数据来源",
    "采集时间",
)
WIDTHS = (38, 20, 28, 10, 28, 12, 52, 16, 42, 18, 22, 20, 27, 27)


def main():
    TARGET.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Messages"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:N1"
    for index, width in enumerate(WIDTHS, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for index in (1, 3, 9, 10, 11, 12):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].number_format = "@"
    workbook.save(TARGET)


if __name__ == "__main__":
    main()
