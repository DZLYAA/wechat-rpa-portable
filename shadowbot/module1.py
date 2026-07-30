import json
import os
from pathlib import Path
from urllib.parse import unquote, urlparse


HEADERS = (
    "消息键",
    "本地时间",
    "会话ID",
    "方向",
    "发送者微信名（账号）",
    "消息类型",
    "消息内容",
    "图片预览",
    "媒体路径",
    "本地消息ID",
    "服务器消息ID",
    "排序序号",
    "数据来源",
    "采集时间",
)

TYPE_NAMES = {
    "text": "文本",
    "image": "图片",
    "video": "视频",
    "voice": "语音",
    "file": "文件",
    "link": "链接",
    "location": "位置",
    "sticker": "表情",
    "system": "系统消息",
    "preview": "会话预览",
}

DIRECTION_NAMES = {
    "in": "接收",
    "incoming": "接收",
    "out": "发送",
    "outgoing": "发送",
    "unknown": "未知",
}


def _result_args(value):
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}


def _package_root():
    locator = (
        Path(os.environ.get("LOCALAPPDATA", str(Path.home())))
        / "WechatRpaPortable"
        / "install.json"
    )
    if locator.is_file():
        data = json.loads(locator.read_text(encoding="utf-8-sig"))
        root = Path(data.get("packageRoot") or "")
        if root.is_dir():
            return root
    bundled = Path(__file__).resolve().parent.parent
    if (bundled / "config" / "app.json").is_file():
        return bundled
    raise RuntimeError("找不到便携包，请先运行 setup.exe 完成配置")


def _expand_path(value, root):
    text = os.path.expandvars(str(value or "").strip())
    path = Path(text).expanduser()
    return path if path.is_absolute() else root / path


def _local_media_path(value):
    value = str(value or "").strip()
    if value.lower().startswith("file:"):
        parsed = urlparse(value)
        value = unquote(parsed.path)
        if len(value) >= 3 and value[0] == "/" and value[2] == ":":
            value = value[1:]
        value = value.replace("/", os.sep)
    return os.path.normpath(value) if value else ""


def _last_row(sheet):
    return max(1, int(sheet.Cells(sheet.Rows.Count, 1).End(-4162).Row))


def _existing_keys(sheet):
    last = _last_row(sheet)
    if last < 2:
        return set()
    values = sheet.Range(sheet.Cells(2, 1), sheet.Cells(last, 1)).Value
    if values is None:
        return set()
    if not isinstance(values, tuple):
        values = ((values,),)
    keys = set()
    for row in values:
        value = row[0] if isinstance(row, tuple) and row else row
        if value not in (None, ""):
            keys.add(str(value).strip())
    return keys


def _ensure_sheet(workbook):
    try:
        sheet = workbook.Worksheets("Messages")
    except Exception:
        sheet = workbook.Worksheets(1)
        sheet.Name = "Messages"
    sheet.Range(sheet.Cells(1, 1), sheet.Cells(1, len(HEADERS))).Value = (HEADERS,)
    return sheet


def _format_sheet(sheet, last_row):
    header = sheet.Range(sheet.Cells(1, 1), sheet.Cells(1, len(HEADERS)))
    header.Font.Bold = True
    header.Font.Color = 0xFFFFFF
    header.Interior.Color = 0x784F1F
    header.HorizontalAlignment = -4108
    header.VerticalAlignment = -4108
    sheet.Rows(1).RowHeight = 28
    widths = (38, 20, 28, 10, 28, 12, 52, 16, 42, 18, 22, 20, 27, 27)
    for index, width in enumerate(widths, 1):
        sheet.Columns(index).ColumnWidth = width
    sheet.Columns(1).NumberFormat = "@"
    sheet.Columns(3).NumberFormat = "@"
    sheet.Columns(9).NumberFormat = "@"
    sheet.Columns(10).NumberFormat = "@"
    sheet.Columns(11).NumberFormat = "@"
    sheet.Columns(12).NumberFormat = "@"
    if last_row >= 2:
        body = sheet.Range(sheet.Cells(2, 1), sheet.Cells(last_row, len(HEADERS)))
        body.VerticalAlignment = -4160
        body.WrapText = True
    if sheet.AutoFilterMode:
        sheet.AutoFilterMode = False
    sheet.Range(sheet.Cells(1, 1), sheet.Cells(max(last_row, 1), len(HEADERS))).AutoFilter()
    window = sheet.Application.ActiveWindow
    if window is not None:
        sheet.Activate()
        window.SplitRow = 1
        window.FreezePanes = True


def _add_media(sheet, start_row, rows, embed_images):
    for offset, message in enumerate(rows):
        excel_row = start_row + offset
        media_path = _local_media_path(message.get("media_path"))
        if not media_path:
            continue
        path_cell = sheet.Cells(excel_row, 9)
        path_cell.Value = media_path
        if os.path.exists(media_path):
            try:
                sheet.Hyperlinks.Add(path_cell, media_path, "", "", "打开原文件")
            except Exception:
                pass
        is_image = str(message.get("type_raw") or "").lower() == "image"
        if not embed_images or not is_image or not os.path.isfile(media_path):
            continue
        try:
            cell = sheet.Cells(excel_row, 8)
            picture = sheet.Shapes.AddPicture(
                media_path, False, True, cell.Left + 2, cell.Top + 2, -1, -1
            )
            picture.LockAspectRatio = True
            if picture.Width > 82:
                picture.Width = 82
            if picture.Height > 82:
                picture.Height = 82
            picture.Left = cell.Left + max(2, (cell.Width - picture.Width) / 2)
            picture.Top = cell.Top + 2
            picture.Placement = 1
            sheet.Rows(excel_row).RowHeight = 88
        except Exception:
            sheet.Cells(excel_row, 8).Value = "预览失败，可点击媒体路径"


def _write_excel(excel_path, messages, embed_images):
    try:
        return _write_excel_com(excel_path, messages, embed_images)
    except Exception as error:
        text = str(error)
        hresult = getattr(error, "hresult", None)
        if hresult is None and error.args and isinstance(error.args[0], int):
            hresult = error.args[0]
        if (
            hresult != -2147221005
            and "无效的类字符串" not in text
            and "Invalid class string" not in text
        ):
            raise
        return _write_excel_openpyxl(excel_path, messages, embed_images)


def _write_excel_com(excel_path, messages, embed_images):
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = None
    try:
        workbook = excel.Workbooks.Open(str(excel_path), UpdateLinks=0, ReadOnly=False)
        if workbook.ReadOnly:
            raise RuntimeError("Excel 文件处于只读或被占用状态")
        sheet = _ensure_sheet(workbook)
        known = _existing_keys(sheet)
        rows = [
            message
            for message in messages
            if str(message.get("message_key") or "").strip() not in known
        ]
        if not rows:
            _format_sheet(sheet, _last_row(sheet))
            workbook.Save()
            return 0

        start_row = _last_row(sheet) + 1
        values = []
        for message in rows:
            raw_type = str(message.get("type_raw") or "")
            raw_direction = str(message.get("direction") or "")
            values.append(
                (
                    str(message.get("message_key") or ""),
                    str(message.get("local_time") or ""),
                    str(message.get("session_id") or ""),
                    DIRECTION_NAMES.get(raw_direction.lower(), raw_direction),
                    str(message.get("sender_username") or ""),
                    TYPE_NAMES.get(raw_type.lower(), raw_type),
                    str(message.get("content") or ""),
                    "",
                    _local_media_path(message.get("media_path")),
                    str(message.get("local_id") or ""),
                    str(message.get("server_id_text") or ""),
                    str(message.get("sort_seq_text") or ""),
                    str(message.get("source") or ""),
                    str(message.get("collected_at") or ""),
                )
            )
        end_row = start_row + len(values) - 1
        sheet.Range(sheet.Cells(start_row, 1), sheet.Cells(end_row, len(HEADERS))).Value = tuple(values)
        _add_media(sheet, start_row, rows, embed_images)
        _format_sheet(sheet, end_row)
        workbook.Save()
        return len(rows)
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        excel.Quit()


def _write_excel_openpyxl(excel_path, messages, embed_images):
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    try:
        workbook = load_workbook(excel_path)
    except PermissionError as error:
        raise RuntimeError("Excel 文件被占用，请关闭工作簿后重试") from error
    sheet = workbook["Messages"] if "Messages" in workbook.sheetnames else workbook.active
    sheet.title = "Messages"
    for index, header in enumerate(HEADERS, 1):
        sheet.cell(1, index).value = header
    known = {
        str(sheet.cell(row, 1).value).strip()
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value not in (None, "")
    }
    rows = [
        message
        for message in messages
        if str(message.get("message_key") or "").strip() not in known
    ]
    for message in rows:
        raw_type = str(message.get("type_raw") or "")
        raw_direction = str(message.get("direction") or "")
        media_path = _local_media_path(message.get("media_path"))
        sheet.append(
            (
                str(message.get("message_key") or ""),
                str(message.get("local_time") or ""),
                str(message.get("session_id") or ""),
                DIRECTION_NAMES.get(raw_direction.lower(), raw_direction),
                str(message.get("sender_username") or ""),
                TYPE_NAMES.get(raw_type.lower(), raw_type),
                str(message.get("content") or ""),
                "",
                media_path,
                str(message.get("local_id") or ""),
                str(message.get("server_id_text") or ""),
                str(message.get("sort_seq_text") or ""),
                str(message.get("source") or ""),
                str(message.get("collected_at") or ""),
            )
        )
        row_number = sheet.max_row
        if media_path and os.path.isfile(media_path):
            path_cell = sheet.cell(row_number, 9)
            path_cell.hyperlink = media_path
            path_cell.style = "Hyperlink"
        if (
            embed_images
            and raw_type.lower() == "image"
            and os.path.isfile(media_path)
        ):
            try:
                picture = Image(media_path)
                picture.thumbnail((82, 82))
                sheet.add_image(picture, f"H{row_number}")
                sheet.row_dimensions[row_number].height = 68
            except Exception:
                sheet.cell(row_number, 8).value = "预览失败，可点击媒体路径"

    for cell in sheet[1]:
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    widths = (38, 20, 28, 10, 28, 12, 52, 16, 42, 18, 22, 20, 27, 27)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index in (1, 3, 9, 10, 11, 12):
        for row_number in range(2, sheet.max_row + 1):
            sheet.cell(row_number, index).number_format = "@"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{max(sheet.max_row, 1)}"
    try:
        workbook.save(excel_path)
    except PermissionError as error:
        raise RuntimeError("Excel 文件被占用，请关闭工作簿后重试") from error
    finally:
        workbook.close()
    return len(rows)


def main(args=None):
    result = _result_args(args)
    result.update(
        {
            "parse_ok": False,
            "excel_saved": False,
            "message_count": 0,
            "parse_error": "",
        }
    )
    try:
        root = _package_root()
        config = json.loads(
            (root / "config" / "app.json").read_text(encoding="utf-8-sig")
        )
        message_file = root / "state" / "new_messages.json"
        payload = json.loads(message_file.read_text(encoding="utf-8-sig"))
        if not payload.get("ok"):
            error = payload.get("error") or {}
            raise RuntimeError(error.get("message") or "采集器返回失败")
        messages = (payload.get("data") or {}).get("messages") or []
        if not isinstance(messages, list):
            raise RuntimeError("new_messages.json 的 messages 不是列表")
        result["parse_ok"] = True
        excel_path = _expand_path(
            config.get("excelPath") or r"data\wechat_messages.xlsx", root
        )
        if not excel_path.is_file():
            raise RuntimeError("找不到 Excel 文件，请重新运行 setup.exe")
        result["message_count"] = _write_excel(
            excel_path, messages, bool(config.get("embedImages", True))
        )
        result["excel_saved"] = True
    except Exception as error:
        result["parse_error"] = str(error)[:1000]
    return result
